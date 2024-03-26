from flask import Flask, request, jsonify
from firebase_admin import credentials, storage
from paddleocr import PaddleOCR
import cv2
import firebase_admin
import csv
from langchain.output_parsers import PydanticOutputParser
from langchain.prompts import PromptTemplate
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_openai import ChatOpenAI
import csv
import numpy as np
from skimage.filters import threshold_local
import openpyxl
from deskew import deskew
from local_constants import OPENAPI_KEY

app = Flask(__name__)

# firebase credentials
cred = credentials.Certificate(
    "/Users/alyshawang/Documents/ReceiptScanner/receiptscanner-7df72-firebase-adminsdk-68x0t-3ec3d2e781.json")
firebase_admin.initialize_app(
    cred, {"storageBucket": "receiptscanner-7df72.appspot.com"})
bucket = storage.bucket()

ocr = PaddleOCR(use_angle_cls=True, use_gpu=False)

openai_api_key = OPENAPI_KEY
model = ChatOpenAI(api_key=openai_api_key, temperature=0)

class ReceiptInfo(BaseModel):
    total_amount: float = Field(
        description="Total amount paid on the receipt")
    date: str = Field(description="Date on the receipt")
    category: str = Field(description="Company tax category")

parser = PydanticOutputParser(pydantic_object=ReceiptInfo)

prompt = PromptTemplate(
    template="Answer the user query.\n{format_instructions}\n{query}\n",
    input_variables=["query"],
    partial_variables={
        "format_instructions": parser.get_format_instructions()},
)

chain = prompt | model | parser


def process_image(image_bytes):
    # image processing
    with open('temp_image.jpg', 'wb') as file:
        file.write(image_bytes)

    frame = cv2.imread('temp_image.jpg')
    deskewed = deskew("temp_image.jpg")
    cv2.imwrite("1-deskewed-opencv.jpg", deskewed)
    framec = cv2.imread('1-deskewed-opencv.jpg')

    gray = cv2.cvtColor(framec, cv2.COLOR_BGR2GRAY)

    binary_adaptive = threshold_local(gray, block_size=101, offset=20)
    binary_adaptive = (gray > binary_adaptive).astype(np.uint8) * 255

    edge_image = cv2.Canny(binary_adaptive, 250, 200)

    cv2.imwrite('edge.jpg', edge_image)
    contours, hierarchy = cv2.findContours(edge_image,
                                           cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    img = cv2.cvtColor(binary_adaptive, cv2.COLOR_GRAY2BGR)

    cv2.drawContours(img, contours, -1, (0, 255, 0), 3)
    cv2.imwrite('processed_image.jpg', img)

    result = ocr.ocr(binary_adaptive, cls=True)

    extracted_text = ' '.join([word[1][0] for line in result for word in line])

    return extracted_text


@app.route('/process_image', methods=['POST'])
def process_image_route():
    # extract text
    print("Full URL:", request.url)

    folder_name = request.args.get('folderName')

    print("Received folder name from Flutter:", folder_name)

    blobs = bucket.list_blobs(prefix=folder_name + '/')
    for blob in blobs:
        image_bytes = blob.download_as_bytes()

        extracted_text = process_image(image_bytes)

        print("Extracted Text:")
        print(extracted_text)

        with open('ocr_results.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([extracted_text])

    ocr_results_file = 'ocr_results.csv'

    # parse result
    with open(ocr_results_file, 'r', newline='') as file:
        reader = csv.reader(file)
        for row in reader:
            receipt_text = row[0]
        company_query = f"{receipt_text}\n Extract the total amount paid after everything and date from the receipt text provided. Return the date in the format MM/DD/YYYY. Also from the receipt company name and products purchased, extract the tax category it belongs to: Groceries, Clothes, Office, Vehicle, Meals, Rent, Phone & Utilities, Supplies, or Assets"
        result = chain.invoke(
            {"query": company_query, "openai_api_key": openai_api_key})

        print("Total Amount Paid:", result.total_amount)
        print("Date:", result.date)
        print("Category:", result.category)
    return jsonify({
        'total_amount': result.total_amount,
        'date': result.date,
        'category': result.category
    }), 200


@app.route('/upload_data', methods=['POST'])
def upload():
    # upload to excel
    output_file = 'receipt_info.xlsx'

    try:
        workbook = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    category_worksheets = {}
    data = request.get_json()
    date = data.get('date')
    category = data.get('category')
    total_amount = data.get('total_amount')
    if date and category and total_amount:
        try:
            if category not in category_worksheets:
                if category in workbook.sheetnames:
                    category_worksheets[category] = workbook[category]
                else:
                    category_worksheets[category] = workbook.create_sheet(
                        title=category)
                    category_worksheets[category].append(
                        ["Date", "Total Amount Paid"])

            category_worksheets[category].append([date, total_amount])
            workbook.save(output_file)

            return jsonify({'message': 'Data saved successfully'}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:
        return jsonify({'error': 'Invalid data received'}), 400


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5001, debug=True)
